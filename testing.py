import openpyxl
import datetime
from docxtpl import DocxTemplate

path = "HeroAcheiver.xlsx"
  
wb = openpyxl.load_workbook(path, data_only = True, read_only = True) 
  
sheet = wb['details']
maxrow = sheet.max_row

print(maxrow)

# sheet.cell(2,1).value = 'Bike Name'
# wb.save('testing.xlsx')    

for ab in range(2, maxrow+1):
    document = DocxTemplate('template.docx')
    bname = sheet.cell(ab, 1).value
    print(bname)
    bprice = sheet.cell(ab, 2).value
    bHelmet = sheet.cell(ab, 3).value
    purchaseExpense = sheet.cell(ab, 4).value
    serviceExpense = sheet.cell(ab, 5).value
    fuelExpense = sheet.cell(ab, 6).value
    totalExpense = sheet.cell(ab, 7).value
    mileage = sheet.cell(ab,8).value
    purchasedDate = sheet.cell(ab,9).value
    purchasedDate = purchasedDate.strftime('%d/%b/%Y')

    context = { 'BNAME' : bname,
                'BPRICE': bprice,
                'HELMET': bHelmet,
                'PURCHASE_EXPENSE': purchaseExpense,
                'SERVICE_EXPENSE': serviceExpense,
                'FUEL_EXPENSE': fuelExpense,
                'TOTAL_EXPENSE': totalExpense,
                'MILEAGE': mileage,
                'PURCHASED_DATE': purchasedDate
                }
    document.render(context)
    document.save('output/'+bname+'.docx')